"""
data_manager.py
Módulo de persistencia de datos.
Gestiona todas las operaciones de lectura y escritura sobre el archivo Excel.
Utiliza openpyxl para manipular hojas de Productos, Zonas y Pedidos.
"""

import openpyxl
from openpyxl import Workbook
from datetime import datetime
from pathlib import Path


def nombre_archivo_excel():
    #Centraliza el nombre para evitar errores de typo entre funciones.
    return "comedor_db.xlsx"


def inicializar_base_de_datos():
    #Crea el archivo Excel con la estructura inicial si no existe.
    #Si el archivo ya existe, no hace nada para preservar pedidos anteriores.

    archivo = nombre_archivo_excel()

    # Verifica si el archivo ya existe usando pathlib.
    if Path(archivo).exists():
        return

    try:
        wb = Workbook()

        # Hoja 1: Productos (menú).
        ws_prod = wb.active
        ws_prod.title = "Productos"
        ws_prod.append(["ID", "Nombre", "Precio", "Stock"])
        ws_prod.append([1, "Hamburguesa Completa", 2500, 20])
        ws_prod.append([2, "Pizza Muzarella", 3000, 15])
        ws_prod.append([3, "Empanadas (docena)", 1800, 10])
        ws_prod.append([4, "Milanesa con Fritas", 2800, 12])
        ws_prod.append([5, "Gaseosa 1.5L", 1200, 30])

        # Hoja 2: Zonas (centregas).
        ws_zonas = wb.create_sheet("Zonas")
        ws_zonas.append(["Nombre", "Entrega", "Tiempo"])
        ws_zonas.append(["Centro", "SI", "30 min"])
        ws_zonas.append(["Alberdi", "SI", "45 min"])
        ws_zonas.append(["Cerro", "SI", "40 min"])
        ws_zonas.append(["Granadero Baigorria", "NO", "Retiro en local"])
        ws_zonas.append(["Funes", "NO", "Retiro en local"])

        # Hoja 3: Pedidos (registro).
        ws_pedidos = wb.create_sheet("Pedidos")
        ws_pedidos.append([
            "Nro_Orden", "Fecha", "Cliente", "Telefono",
            "Direccion", "Producto", "Cantidad", "Total", "Estado", "Zona"
        ])

        wb.save(archivo)
        print("✅ Base de datos inicializada correctamente.")
    except Exception as e:
        raise Exception(f"Error al inicializar base de datos: {e}")


def obtener_productos():
    #Lee la hoja "Productos" del Excel y devuelve lista de diccionarios.

    archivo = nombre_archivo_excel()
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb["Productos"]
        productos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                productos.append({
                    "id": row[0],
                    "nombre": row[1],
                    "precio": row[2],
                    "stock": row[3]
                })
        wb.close()
        return productos
    except Exception as e:
        raise Exception(f"Error al leer productos: {e}")


def obtener_zonas():
    #Lee la hoja "Zonas" del Excel y devuelve lista de diccionarios.

    archivo = nombre_archivo_excel()
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb["Zonas"]
        zonas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                zonas.append({
                    "nombre": row[0],
                    "entrega": row[1],
                    "tiempo": row[2]
                })
        wb.close()
        return zonas
    except Exception as e:
        raise Exception(f"Error al leer zonas: {e}")


def obtener_siguiente_nro_orden():
    #Calcula el siguiente número de orden, leyendo la última fila de la hoja "Pedidos".
    archivo = nombre_archivo_excel()
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb["Pedidos"]
        nro = 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                nro = row[0] + 1
        wb.close()
        return nro
    except Exception as e:
        raise Exception(f"Error al obtener número de orden: {e}")


def guardar_pedido(nro_orden, cliente, telefono, direccion, producto, cantidad, total, estado, zona):
    #Registra un nuevo pedido en la hoja "Pedidos" del Excel.

    archivo = nombre_archivo_excel()
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb["Pedidos"]
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws.append([
            nro_orden, fecha, cliente, telefono,
            direccion, producto, cantidad, total, estado, zona
        ])
        wb.save(archivo)
        wb.close()
    except Exception as e:
        raise Exception(f"Error al guardar pedido: {e}")


def actualizar_stock(producto_id, cantidad_comprada):
    #Descuenta del stock la cantidad comprada en la hoja "Productos".

    archivo = nombre_archivo_excel()
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb["Productos"]
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == producto_id:
                stock_actual = row[3].value
                # Descuenta la cantidad comprada del stock actual.
                row[3].value = stock_actual - cantidad_comprada
                break
        wb.save(archivo)
        wb.close()
    except Exception as e:
        raise Exception(f"Error al actualizar stock: {e}")


def obtener_info_zona(nombre_zona, zonas):
    #Busca una zona por nombre dentro de la lista de zonas cargada.

    for z in zonas:
        if z["nombre"].upper() == nombre_zona.upper():
            return z
    raise ValueError("Zona no encontrada.")
